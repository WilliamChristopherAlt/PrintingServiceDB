#!/usr/bin/env python3
"""
Extract all content from Students.xlsx to a text file
"""

import openpyxl
from pathlib import Path

EXCEL_FILE = Path(__file__).parent / "Students.xlsx"
OUTPUT_FILE = Path(__file__).parent / "students_excel_content.txt"

def extract_all_content():
    """Extract all content from all sheets in the Excel file."""
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    output_lines = []
    output_lines.append("=" * 80)
    output_lines.append("EXTRACTED CONTENT FROM STUDENTS.XLSX")
    output_lines.append("=" * 80)
    output_lines.append("")
    
    for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        
        output_lines.append("=" * 80)
        output_lines.append(f"SHEET {sheet_idx}: {sheet_name}")
        output_lines.append("=" * 80)
        output_lines.append(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
        output_lines.append("")
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append("")
        
        # Write headers
        output_lines.append("HEADERS:")
        output_lines.append(" | ".join(headers))
        output_lines.append("")
        output_lines.append("-" * 80)
        output_lines.append("")
        
        # Extract all rows
        row_count = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # Skip completely empty rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            
            row_count += 1
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append("")
                else:
                    # Convert to string and handle newlines
                    cell_str = str(cell).strip()
                    # Replace newlines with spaces for better readability
                    cell_str = cell_str.replace('\n', ' ').replace('\r', ' ')
                    row_data.append(cell_str)
            
            output_lines.append(f"ROW {row_idx}:")
            output_lines.append(" | ".join(row_data))
            output_lines.append("")
        
        output_lines.append(f"Total non-empty rows: {row_count}")
        output_lines.append("")
        output_lines.append("")
    
    # Summary
    output_lines.append("=" * 80)
    output_lines.append("SUMMARY")
    output_lines.append("=" * 80)
    output_lines.append(f"Total sheets: {len(wb.sheetnames)}")
    output_lines.append(f"Sheet names: {', '.join(wb.sheetnames)}")
    output_lines.append("")
    
    # Write to file
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(output_lines))
    
    print(f"Extraction complete!")
    print(f"Output written to: {OUTPUT_FILE}")
    print(f"Total sheets processed: {len(wb.sheetnames)}")
    
    return output_lines

if __name__ == "__main__":
    extract_all_content()

