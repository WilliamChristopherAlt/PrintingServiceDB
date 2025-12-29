#!/usr/bin/env python3
"""
Extract test cases from Students.xlsx and generate content for section 6.1
"""

import openpyxl
from pathlib import Path
from collections import defaultdict

EXCEL_FILE = Path(__file__).parent / "Students.xlsx"
OUTPUT_FILE = Path(__file__).parent / "test_cases_output.txt"

def extract_test_cases():
    """Extract test cases from Excel and generate formatted output."""
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    output_lines = []
    output_lines.append("=" * 80)
    output_lines.append("TEST CASES EXTRACTED FROM STUDENTS.XLSX")
    output_lines.append("=" * 80)
    output_lines.append("")
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        output_lines.append(f"\n{'=' * 80}")
        output_lines.append(f"SHEET: {sheet_name}")
        output_lines.append(f"{'=' * 80}")
        output_lines.append(f"Total rows: {ws.max_row}, Columns: {ws.max_column}")
        output_lines.append("")
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append("")
        
        output_lines.append(f"Headers: {headers[:10]}")
        output_lines.append("")
        
        # Extract test cases (skip header row)
        test_cases = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0] or str(row[0]).strip() == "":
                continue
            
            test_case = {
                'id': str(row[0]).strip() if row[0] else "",
                'name': str(row[1]).strip() if row[1] else "",
                'priority': str(row[2]).strip() if row[2] else "",
                'steps': str(row[3]).strip() if row[3] else "",
                'expected': str(row[4]).strip() if row[4] else "",
                'result': str(row[5]).strip() if row[5] else "",
            }
            
            if test_case['id']:
                test_cases.append(test_case)
        
        # Group by priority
        by_priority = defaultdict(list)
        for tc in test_cases:
            priority = tc['priority'] if tc['priority'] else "Không xác định"
            by_priority[priority].append(tc)
        
        output_lines.append(f"Total test cases: {len(test_cases)}")
        output_lines.append(f"By priority: {dict(by_priority)}")
        output_lines.append("")
        
        # Count results
        result_counts = defaultdict(int)
        for tc in test_cases:
            result = tc['result'] if tc['result'] else "Chưa test"
            result_counts[result] += 1
        
        output_lines.append(f"Results: {dict(result_counts)}")
        output_lines.append("")
        
        # Sample first 3 test cases
        output_lines.append("Sample test cases (first 3):")
        for i, tc in enumerate(test_cases[:3], 1):
            output_lines.append(f"\n{i}. {tc['id']}: {tc['name']}")
            output_lines.append(f"   Priority: {tc['priority']}")
            output_lines.append(f"   Steps: {tc['steps'][:100]}...")
            output_lines.append(f"   Expected: {tc['expected'][:100]}...")
            output_lines.append(f"   Result: {tc['result']}")
    
    # Write to file
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(output_lines))
    
    print(f"Extraction complete! Output written to: {OUTPUT_FILE}")
    return test_cases, wb.sheetnames

if __name__ == "__main__":
    extract_test_cases()

