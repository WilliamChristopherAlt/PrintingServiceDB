#!/usr/bin/env python3
"""
Test script to examine the structure of Students.xlsx file
"""

import openpyxl
from pathlib import Path

EXCEL_FILE = Path(__file__).parent / "Students.xlsx"

def examine_excel():
    """Examine the structure of the Excel file."""
    print("=" * 60)
    print("EXAMINING EXCEL FILE STRUCTURE")
    print("=" * 60)
    
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    print(f"\nTotal sheets: {len(wb.sheetnames)}")
    print(f"Sheet names: {wb.sheetnames}")
    print("\n" + "=" * 60)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
        print("\nFirst 10 rows:")
        print("-" * 60)
        
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i > 10:
                break
            # Filter out None values for cleaner output
            row_data = [str(cell) if cell is not None else "" for cell in row[:15]]
            print(f"Row {i:2d}: {' | '.join(row_data)}")
        
        print("\n" + "=" * 60)

if __name__ == "__main__":
    examine_excel()

