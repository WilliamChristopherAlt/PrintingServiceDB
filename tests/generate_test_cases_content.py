#!/usr/bin/env python3
"""
Generate test cases content for section 6.1 from Students.xlsx
"""

import openpyxl
from pathlib import Path
from collections import defaultdict

EXCEL_FILE = Path(__file__).parent / "Students.xlsx"
OUTPUT_FILE = Path(__file__).parent / "test_cases_content.txt"

# Map sheet names to use case names
SHEET_TO_USE_CASE = {
    "SV-1-Thành phần chung": "UC-01: Thành phần chung",
    "SV-2-Trang chủ": "UC-02: Trang chủ",
    "SV-3-Trang nạp tiền tài khoản": "UC-03: Nạp tiền tài khoản",
    "SV-4-Trang in tài liệu": "UC-04: In tài liệu",
    "SV-5-Trang lịch sử in": "UC-05: Lịch sử in",
    "SV-6-Thông tin sinh viên": "UC-06: Thông tin sinh viên",
}

def extract_and_analyze():
    """Extract test cases and generate formatted content."""
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    output_lines = []
    output_lines.append("6.1. Các Test Case")
    output_lines.append("")
    output_lines.append("Hệ thống đã được kiểm thử kỹ lưỡng với tổng cộng các test case được thiết kế cho các màn hình chính của sinh viên. Các test case được phân loại theo mức độ ưu tiên (Cao, Trung bình) và được thực hiện để đảm bảo chất lượng và tính ổn định của hệ thống.")
    output_lines.append("")
    
    total_test_cases = 0
    total_passed = 0
    total_failed = 0
    total_skipped = 0
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_TO_USE_CASE:
            continue
            
        use_case_name = SHEET_TO_USE_CASE[sheet_name]
        ws = wb[sheet_name]
        
        # Extract test cases
        test_cases = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0] or str(row[0]).strip() == "":
                continue
            
            test_case = {
                'id': str(row[0]).strip() if row[0] else "",
                'name': str(row[1]).strip() if row[1] else "",
                'priority': str(row[2]).strip() if row[2] else "",
                'steps': str(row[3]).strip() if row[3] else "",
                'expected': str(row[4]).strip() if row[4] else "",
                'result': str(row[5]).strip() if row[5] else "",
            }
            
            if test_case['id']:
                test_cases.append(test_case)
        
        if not test_cases:
            continue
        
        # Count results
        passed = sum(1 for tc in test_cases if tc['result'].upper() == 'PASS')
        failed = sum(1 for tc in test_cases if tc['result'].upper() == 'FAIL')
        skipped = sum(1 for tc in test_cases if tc['result'].upper() in ['SKIP', 'SKIPPED', ''])
        not_tested = sum(1 for tc in test_cases if not tc['result'] or tc['result'].strip() == '')
        
        total_test_cases += len(test_cases)
        total_passed += passed
        total_failed += failed
        total_skipped += skipped
        
        # Group by priority
        by_priority = defaultdict(list)
        for tc in test_cases:
            priority = tc['priority'] if tc['priority'] else "Không xác định"
            by_priority[priority].append(tc)
        
        # Write section for this use case
        output_lines.append(f"6.1.{len([s for s in wb.sheetnames if s in SHEET_TO_USE_CASE and wb.sheetnames.index(s) < wb.sheetnames.index(sheet_name)]) + 1}. {use_case_name}")
        output_lines.append("")
        output_lines.append(f"Màn hình {sheet_name} được kiểm thử với {len(test_cases)} test case, trong đó {passed} test case đạt (Pass), {failed} test case thất bại (Fail), và {skipped + not_tested} test case chưa được thực hiện hoặc bị bỏ qua.")
        output_lines.append("")
        
        # Summary by priority
        if 'Cao' in by_priority:
            high_priority = by_priority['Cao']
            high_passed = sum(1 for tc in high_priority if tc['result'].upper() == 'PASS')
            output_lines.append(f"Trong số {len(high_priority)} test case mức độ cao, {high_passed} test case đạt yêu cầu.")
            output_lines.append("")
        
        if 'TB' in by_priority:
            medium_priority = by_priority['TB']
            medium_passed = sum(1 for tc in medium_priority if tc['result'].upper() == 'PASS')
            output_lines.append(f"Trong số {len(medium_priority)} test case mức độ trung bình, {medium_passed} test case đạt yêu cầu.")
            output_lines.append("")
        
        # Placeholder for table
        output_lines.append(f"[Bảng 6.{len([s for s in wb.sheetnames if s in SHEET_TO_USE_CASE and wb.sheetnames.index(s) < wb.sheetnames.index(sheet_name)]) + 1}: Các test case cho {use_case_name}]")
        output_lines.append("")
        output_lines.append("Bảng trên liệt kê các test case chính cho use case này, bao gồm ID test case, tên test case, mức độ ưu tiên, các bước thực hiện, kết quả mong đợi và kết quả thực tế.")
        output_lines.append("")
    
    # Overall summary
    output_lines.append("6.1.7. Tổng kết")
    output_lines.append("")
    output_lines.append(f"Tổng cộng, hệ thống đã được kiểm thử với {total_test_cases} test case trên {len([s for s in wb.sheetnames if s in SHEET_TO_USE_CASE])} màn hình chính. Kết quả tổng thể:")
    output_lines.append("")
    output_lines.append(f"• Số test case đạt: {total_passed} ({total_passed/total_test_cases*100:.1f}%)")
    output_lines.append(f"• Số test case thất bại: {total_failed} ({total_failed/total_test_cases*100:.1f}%)")
    output_lines.append(f"• Số test case chưa thực hiện/bỏ qua: {total_skipped} ({total_skipped/total_test_cases*100:.1f}%)")
    output_lines.append("")
    output_lines.append("Tỷ lệ test case đạt yêu cầu cho thấy hệ thống đã được kiểm thử kỹ lưỡng và đáp ứng các yêu cầu chức năng cơ bản. Các test case thất bại và chưa thực hiện sẽ được xử lý trong các phiên bản tiếp theo.")
    output_lines.append("")
    
    # Write to file
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(output_lines))
    
    print(f"Content generated! Output written to: {OUTPUT_FILE}")
    print(f"Total test cases: {total_test_cases}")
    print(f"Passed: {total_passed}, Failed: {total_failed}, Skipped: {total_skipped}")
    
    return output_lines

if __name__ == "__main__":
    extract_and_analyze()

